from io import BytesIO
from django.template.loader import get_template
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def generate_pdf(savings, loans):
    template = get_template('reports/pdf_template.html')  # Create a template for PDF
    context = {'savings': savings}
    html = template.render(context)
    pdf_file = HTML(string=html).write_pdf()
    return pdf_file

def generate_excel(savings):
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers
    headers = [
        'Month', 'Owner', 'Received', 'Savings', 'Divine Touch', 'RSS', 'Shares', 
        'Loan', 'Interest', 'Commod', 'Savings Balance', 'Divine Touch Balance', 
        'Rss Balance', 'Shares Balance', 'Loan Balance', 'Interest Balance', 'Commodity Balance',
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}1'] = header

    # Write data
    for row_num, saving in enumerate(savings, start=2):
        worksheet[f'A{row_num}'] = saving.get_month_display()
        worksheet[f'B{row_num}'] = f'{saving.owner.first_name} {saving.owner.last_name}'
        worksheet[f'C{row_num}'] = saving.received
        worksheet[f'D{row_num}'] = saving.savings
        worksheet[f'E{row_num}'] = saving.divine_touch
        worksheet[f'F{row_num}'] = saving.rss
        worksheet[f'G{row_num}'] = saving.share
        worksheet[f'H{row_num}'] = saving.loan
        worksheet[f'I{row_num}'] = saving.interest
        worksheet[f'J{row_num}'] = saving.commod
        worksheet[f'K{row_num}'] = saving.savings_balance
        worksheet[f'L{row_num}'] = saving.divine_touch_balance
        worksheet[f'M{row_num}'] = saving.rss_balance
        worksheet[f'N{row_num}'] = saving.share_balance
        worksheet[f'O{row_num}'] = saving.loan_balance
        worksheet[f'P{row_num}'] = saving.interest_balance
        worksheet[f'Q{row_num}'] = saving.commodity_balance

    workbook.save(output)
    output.seek(0)
    return output.getvalue()